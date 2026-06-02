from __future__ import annotations

import argparse
import json
import math
import shutil
from datetime import date, datetime
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.utils.datetime import from_excel


def arg_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(description="Build a static GitHub Pages workbook viewer.")
    parser.add_argument("--source", required=True, help="Input XLSX file")
    parser.add_argument("--out", required=True, help="Output repository directory")
    return parser


def json_safe(value):
    if isinstance(value, datetime):
        return value.isoformat(sep=" ", timespec="minutes")
    if isinstance(value, date):
        return value.isoformat()
    if isinstance(value, float):
        if math.isnan(value) or math.isinf(value):
            return None
    return value


def color_from_cell(cell, attr: str) -> str | None:
    color = getattr(cell.font if attr == "font" else cell.fill, "color", None)
    if color is None:
        return None
    if color.type == "rgb" and color.rgb:
        value = color.rgb[-6:]
        if value.upper() != "000000":
            return f"#{value}"
    return None


def fill_color(cell) -> str | None:
    fill: PatternFill = cell.fill
    if not fill or fill.patternType in (None, "none"):
        return None
    color = fill.fgColor
    if color and color.type == "rgb" and color.rgb:
        value = color.rgb[-6:]
        if value.upper() not in ("000000", "FFFFFF"):
            return f"#{value}"
    return None


def is_probably_date(cell, value) -> bool:
    if isinstance(value, (datetime, date)):
        return True
    fmt = (cell.number_format or "").lower()
    return isinstance(value, (int, float)) and any(token in fmt for token in ("yy", "mm", "dd")) and not any(
        token in fmt for token in ("0", "#")
    )


def display_value(cell, value):
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.strftime("%Y-%m-%d %H:%M") if value.time().hour or value.time().minute else value.strftime("%Y-%m-%d")
    if isinstance(value, date):
        return value.strftime("%Y-%m-%d")
    if is_probably_date(cell, value):
        try:
            converted = from_excel(value)
            return converted.strftime("%Y-%m-%d")
        except Exception:
            pass
    if isinstance(value, float):
        fmt = cell.number_format or ""
        if "%" in fmt:
            decimals = 2 if ".00" in fmt else 1 if ".0" in fmt else 0
            return f"{value * 100:.{decimals}f}%"
        if value.is_integer():
            return str(int(value))
        return f"{value:.4f}".rstrip("0").rstrip(".")
    return str(value)


def cell_payload(value_cell, formula_cell):
    raw = value_cell.value
    formula = formula_cell.value if formula_cell.data_type == "f" else None
    text = display_value(value_cell, raw)
    style = {}
    fill = fill_color(formula_cell)
    font_color = color_from_cell(formula_cell, "font")
    if fill:
        style["fill"] = fill
    if font_color:
        style["color"] = font_color
    if formula_cell.font and formula_cell.font.bold:
        style["bold"] = True
    if formula_cell.font and formula_cell.font.italic:
        style["italic"] = True
    if formula_cell.alignment:
        if formula_cell.alignment.horizontal:
            style["align"] = formula_cell.alignment.horizontal
        if formula_cell.alignment.vertical:
            style["valign"] = formula_cell.alignment.vertical
    payload = {"text": text}
    if raw not in (None, ""):
        payload["value"] = json_safe(raw)
    if formula:
        payload["formula"] = formula
    if style:
        payload["style"] = style
    return payload


def trim_dimensions(value_ws, formula_ws):
    max_row = value_ws.max_row or 1
    max_col = value_ws.max_column or 1
    last_row = 1
    last_col = 1
    for row in range(1, max_row + 1):
        has_value = False
        for col in range(1, max_col + 1):
            vc = value_ws.cell(row=row, column=col).value
            fc = formula_ws.cell(row=row, column=col).value
            if vc not in (None, "") or fc not in (None, ""):
                has_value = True
                last_col = max(last_col, col)
        if has_value:
            last_row = row
    return last_row, last_col


def workbook_to_json(source: Path):
    wb_values = load_workbook(source, data_only=True)
    wb_formulas = load_workbook(source, data_only=False)
    sheets = []
    for ws_values in wb_values.worksheets:
        ws_formulas = wb_formulas[ws_values.title]
        last_row, last_col = trim_dimensions(ws_values, ws_formulas)
        rows = []
        for row in range(1, last_row + 1):
            row_values = []
            for col in range(1, last_col + 1):
                row_values.append(cell_payload(ws_values.cell(row=row, column=col), ws_formulas.cell(row=row, column=col)))
            rows.append(row_values)
        merges = [str(rng) for rng in ws_formulas.merged_cells.ranges]
        widths = []
        for col in range(1, last_col + 1):
            letter = get_column_letter(col)
            width = ws_formulas.column_dimensions[letter].width
            widths.append(width if width else None)
        sheets.append(
            {
                "name": ws_values.title,
                "rowCount": last_row,
                "colCount": last_col,
                "columns": [get_column_letter(i) for i in range(1, last_col + 1)],
                "widths": widths,
                "merges": merges,
                "rows": rows,
            }
        )
    return {
        "title": source.stem,
        "fileName": source.name,
        "generatedAt": datetime.now().isoformat(timespec="seconds"),
        "sheets": sheets,
    }


INDEX_HTML = """<!doctype html>
<html lang="zh-CN">
<head>
  <meta charset="utf-8" />
  <meta name="viewport" content="width=device-width, initial-scale=1" />
  <title>2026年进口货物物流跟踪表</title>
  <link rel="stylesheet" href="styles.css" />
</head>
<body>
  <header class="app-header">
    <div>
      <p class="eyebrow">Import Logistics Tracker</p>
      <h1 id="workbook-title">2026年进口货物物流跟踪表</h1>
      <p id="meta" class="meta"></p>
    </div>
    <a class="download" href="assets/2026年进口货物物流跟踪表-铁矿协作一部&新拓.xlsx">下载原表</a>
  </header>

  <main>
    <section class="toolbar" aria-label="Workbook controls">
      <nav id="tabs" class="tabs" aria-label="Sheets"></nav>
      <div class="search-wrap">
        <input id="search" type="search" placeholder="搜索当前工作表" autocomplete="off" />
        <span id="count" class="count"></span>
      </div>
    </section>
    <section class="table-shell">
      <table id="sheet-table"></table>
    </section>
  </main>

  <script src="data/workbook.js"></script>
  <script src="app.js"></script>
</body>
</html>
"""


APP_JS = r"""const state = {
  sheetIndex: 0,
  query: "",
};

const workbook = window.WORKBOOK_DATA;
const tabs = document.getElementById("tabs");
const table = document.getElementById("sheet-table");
const search = document.getElementById("search");
const count = document.getElementById("count");

document.getElementById("workbook-title").textContent = workbook.title;
document.getElementById("meta").textContent = `${workbook.sheets.length} 个工作表 · ${new Date(workbook.generatedAt).toLocaleString("zh-CN")}`;

function escapeText(value) {
  return String(value ?? "")
    .replaceAll("&", "&amp;")
    .replaceAll("<", "&lt;")
    .replaceAll(">", "&gt;")
    .replaceAll('"', "&quot;");
}

function styleAttr(style) {
  if (!style) return "";
  const rules = [];
  if (style.fill) rules.push(`background:${style.fill}`);
  if (style.color) rules.push(`color:${style.color}`);
  if (style.bold) rules.push("font-weight:700");
  if (style.italic) rules.push("font-style:italic");
  if (style.align) rules.push(`text-align:${style.align}`);
  if (style.valign) rules.push(`vertical-align:${style.valign}`);
  return rules.length ? ` style="${rules.join(";")}"` : "";
}

function renderTabs() {
  tabs.innerHTML = workbook.sheets.map((sheet, index) => {
    const active = index === state.sheetIndex ? " active" : "";
    return `<button class="tab${active}" data-index="${index}" type="button">${escapeText(sheet.name)}</button>`;
  }).join("");
}

function rowMatches(row, query) {
  if (!query) return true;
  return row.some(cell => String(cell.text ?? cell.value ?? "").toLowerCase().includes(query));
}

function renderTable() {
  const sheet = workbook.sheets[state.sheetIndex];
  const query = state.query.trim().toLowerCase();
  const rows = sheet.rows
    .map((row, index) => ({ row, index }))
    .filter(({ row }) => rowMatches(row, query));

  const colHeaders = ["<th class=\"corner\"></th>", ...sheet.columns.map((col, index) => {
    const width = sheet.widths[index] ? Math.min(Math.max(sheet.widths[index] * 8, 72), 220) : 120;
    return `<th class="col-head" style="min-width:${width}px">${col}</th>`;
  })].join("");

  const body = rows.map(({ row, index }) => {
    const cells = row.map(cell => {
      const classes = ["cell"];
      if (!cell.text) classes.push("empty");
      const title = cell.formula ? ` title="${escapeText(cell.formula)}"` : "";
      return `<td class="${classes.join(" ")}"${title}${styleAttr(cell.style)}>${escapeText(cell.text)}</td>`;
    }).join("");
    return `<tr><th class="row-head">${index + 1}</th>${cells}</tr>`;
  }).join("");

  table.innerHTML = `<thead><tr>${colHeaders}</tr></thead><tbody>${body}</tbody>`;
  count.textContent = `${rows.length} / ${sheet.rowCount} 行`;
}

function render() {
  renderTabs();
  renderTable();
}

tabs.addEventListener("click", event => {
  const button = event.target.closest("button[data-index]");
  if (!button) return;
  state.sheetIndex = Number(button.dataset.index);
  state.query = "";
  search.value = "";
  render();
});

search.addEventListener("input", event => {
  state.query = event.target.value;
  renderTable();
});

render();
"""


STYLES_CSS = """* {
  box-sizing: border-box;
}

:root {
  color-scheme: light;
  --ink: #17212b;
  --muted: #617080;
  --line: #d7dee8;
  --panel: #ffffff;
  --page: #f5f7fa;
  --accent: #0f766e;
  --accent-strong: #0b5f59;
  --header: #edf4f3;
}

body {
  margin: 0;
  min-height: 100vh;
  font-family: "Segoe UI", Arial, "Microsoft YaHei", sans-serif;
  color: var(--ink);
  background: var(--page);
}

.app-header {
  display: flex;
  align-items: flex-end;
  justify-content: space-between;
  gap: 24px;
  padding: 22px 28px 18px;
  border-bottom: 1px solid var(--line);
  background: var(--panel);
}

.eyebrow {
  margin: 0 0 6px;
  color: var(--accent);
  font-size: 12px;
  font-weight: 700;
  letter-spacing: 0;
  text-transform: uppercase;
}

h1 {
  margin: 0;
  font-size: 24px;
  line-height: 1.25;
  font-weight: 700;
}

.meta {
  margin: 7px 0 0;
  color: var(--muted);
  font-size: 13px;
}

.download {
  flex: 0 0 auto;
  border: 1px solid var(--accent);
  border-radius: 6px;
  padding: 9px 13px;
  color: #ffffff;
  background: var(--accent);
  text-decoration: none;
  font-size: 14px;
  font-weight: 600;
}

.download:hover {
  background: var(--accent-strong);
}

main {
  padding: 18px 20px 24px;
}

.toolbar {
  display: flex;
  align-items: center;
  justify-content: space-between;
  gap: 16px;
  margin-bottom: 12px;
}

.tabs {
  display: flex;
  gap: 8px;
  overflow-x: auto;
  padding-bottom: 2px;
}

.tab {
  flex: 0 0 auto;
  border: 1px solid var(--line);
  border-radius: 6px;
  padding: 8px 12px;
  background: #ffffff;
  color: var(--ink);
  font: inherit;
  font-size: 14px;
  cursor: pointer;
}

.tab.active {
  border-color: var(--accent);
  background: var(--header);
  color: var(--accent-strong);
  font-weight: 700;
}

.search-wrap {
  display: flex;
  align-items: center;
  gap: 10px;
  flex: 0 0 360px;
}

input[type="search"] {
  width: 100%;
  min-width: 160px;
  border: 1px solid var(--line);
  border-radius: 6px;
  padding: 9px 11px;
  background: #ffffff;
  color: var(--ink);
  font: inherit;
  font-size: 14px;
}

.count {
  min-width: 80px;
  color: var(--muted);
  font-size: 13px;
  text-align: right;
  white-space: nowrap;
}

.table-shell {
  height: calc(100vh - 174px);
  overflow: auto;
  border: 1px solid var(--line);
  background: #ffffff;
}

table {
  border-collapse: separate;
  border-spacing: 0;
  width: max-content;
  min-width: 100%;
  font-size: 13px;
}

th,
td {
  border-right: 1px solid var(--line);
  border-bottom: 1px solid var(--line);
  padding: 7px 9px;
  max-width: 340px;
  min-height: 32px;
  white-space: pre-wrap;
  overflow-wrap: anywhere;
  line-height: 1.35;
}

thead th {
  position: sticky;
  top: 0;
  z-index: 3;
  background: #e7eef6;
  color: #435364;
  font-weight: 700;
  text-align: center;
}

.corner,
.row-head {
  position: sticky;
  left: 0;
  z-index: 2;
  width: 54px;
  min-width: 54px;
  max-width: 54px;
  background: #eef2f6;
  color: #667586;
  text-align: right;
  font-weight: 600;
}

.corner {
  z-index: 4;
  top: 0;
}

.cell.empty {
  color: transparent;
}

tbody tr:hover td,
tbody tr:hover .row-head {
  background-color: #fff8dc;
}

@media (max-width: 760px) {
  .app-header,
  .toolbar {
    align-items: stretch;
    flex-direction: column;
  }

  .download,
  .search-wrap {
    width: 100%;
  }

  main {
    padding: 12px;
  }

  .table-shell {
    height: calc(100vh - 258px);
  }
}
"""


README_MD = """# 2026年进口货物物流跟踪表

这个仓库把 Excel 工作簿转换成一个可用 GitHub Pages 公开访问的静态页面。

## 文件

- `index.html`：网页入口
- `data/workbook.js`：从 Excel 提取的工作簿数据
- `assets/2026年进口货物物流跟踪表-铁矿协作一部&新拓.xlsx`：原始 Excel 文件
- `scripts/build_logistics_repo.py`：重新生成站点的脚本

## 发布到 GitHub Pages

创建 GitHub 公开仓库后，在本目录执行：

```powershell
git remote add origin https://github.com/<your-user>/<repo-name>.git
git branch -M main
git push -u origin main
```

然后在 GitHub 仓库页面打开 `Settings -> Pages`，选择 `Deploy from a branch`，分支选 `main`，目录选 `/root`。
公开链接会类似：

```text
https://<your-user>.github.io/<repo-name>/
```

注意：公开仓库和 GitHub Pages 会让仓库里的表格数据被互联网上任何人访问。
"""


def write_repo(source: Path, out_dir: Path) -> None:
    if out_dir.exists():
        shutil.rmtree(out_dir)
    (out_dir / "assets").mkdir(parents=True, exist_ok=True)
    (out_dir / "data").mkdir(parents=True, exist_ok=True)
    (out_dir / "scripts").mkdir(parents=True, exist_ok=True)

    data = workbook_to_json(source)
    shutil.copy2(source, out_dir / "assets" / source.name)
    shutil.copy2(Path(__file__), out_dir / "scripts" / Path(__file__).name)
    (out_dir / "index.html").write_text(INDEX_HTML, encoding="utf-8")
    (out_dir / "app.js").write_text(APP_JS, encoding="utf-8")
    (out_dir / "styles.css").write_text(STYLES_CSS, encoding="utf-8")
    (out_dir / "README.md").write_text(README_MD, encoding="utf-8")
    (out_dir / ".nojekyll").write_text("", encoding="utf-8")
    (out_dir / ".gitignore").write_text("~$*.xlsx\n.DS_Store\nThumbs.db\n", encoding="utf-8")
    json_text = json.dumps(data, ensure_ascii=False, separators=(",", ":"))
    (out_dir / "data" / "workbook.js").write_text(f"window.WORKBOOK_DATA={json_text};\n", encoding="utf-8")


def main() -> None:
    args = arg_parser().parse_args()
    write_repo(Path(args.source), Path(args.out))


if __name__ == "__main__":
    main()
